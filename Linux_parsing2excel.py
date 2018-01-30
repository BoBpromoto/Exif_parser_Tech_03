#!/usr/bin/python
# Best of The Best 6th Digital Forensic
# @Author L3ad0xFF

from PIL import Image
import piexif
import sys
import os
import os.path
import openpyxl
from openpyxl import Workbook

f_list = list()
exif_list = list()
tag_value_temp = list()
exif_value_list = list()
split_exif = list()

collect_file_tag = {}
div_tag = {}
div_tag_temp = {}
div_tag_temp_Exif = {}
div_tag_temp_GPS = {}
div_tag_temp_Interop = {}
div_tag_temp_1st = {}

path = os.getcwd()
for files in os.listdir(path) :
	if (files.split('.')[-1] =='jpg') | (files.split('.')[-1] == 'JPG') | (files.split('.')[-1] == 'jpeg') | (files.split('.')[-1] == 'JPEG') :
		f_list.append(files)

print ("\nTotal file : " + ", ".join(f_list))


for filename in f_list :
	exif_dict = piexif.load(filename)
	collect_file_tag[filename] = exif_dict

for ifd in ("0th", "Exif", "GPS", "Interop", "1st") :
	for tag in piexif.TAGS[ifd] :
		for filename in f_list :
			for file_tag in collect_file_tag[filename][ifd] :
				if tag == file_tag :
					exif_list.append(ifd + " : "+ piexif.TAGS[ifd][file_tag]["name"] + " : " +filename+" : "
						+str(collect_file_tag[filename][ifd][file_tag]))
					break

for lenth in range (0, (len(exif_list))) :
	split_exif.append(exif_list[lenth].split(" : "))


#print (split_exif) 
#split_ exif - [0] : ifd / [1] : tag_name / [2] : filename / [3] : value

_0th_list = list()
Exif_list = list()
GPS_list = list()
Interop_list = list()
_1st_list = list()

_0th_file_split = list()
Exif_file_split = list()
GPS__file_split = list()
Interop_file_split = list()
_1st_file_split = list()

_0th_tag_list = list()
Exif_tag_list = list()
GPS_tag_list = list()
Interop_tag_list = list()
_1st_tag_list = list()

_0th_value_list = list()
Exif_value_list = list()
GPS_value_list = list()
Interop_value_list = list()
_1st_value_list = list()
final_tag_value = {}

for lenth in range (0, len(split_exif)) :
	if  "0th"== split_exif[lenth][0] :
		_0th_list.append(split_exif[lenth][1:4])
		_0th_tag_list.append(split_exif[lenth][1])
		_0th_file_split.append(split_exif[lenth][2])
		_0th_value_list.append(split_exif[lenth][3])
	elif  "Exif"== split_exif[lenth][0] :
		Exif_list.append(split_exif[lenth][1:4])
		Exif_tag_list.append(split_exif[lenth][1])
		Exif_file_split.append(split_exif[lenth][2])
		Exif_value_list.append(split_exif[lenth][3])
	elif  "GPS"== split_exif[lenth][0] :
		GPS_list.append(split_exif[lenth][1:4])
		GPS_tag_list.append(split_exif[lenth][1])
		GPS__file_split.append(split_exif[lenth][2])
		GPS_value_list.append(split_exif[lenth][3])
	elif  "Interop"== split_exif[lenth][0] :
		Interop_list.append(split_exif[lenth][1:4])
		Interop_tag_list.append(split_exif[lenth][1])
		Interop_file_split.append(split_exif[lenth][2])
		Interop_value_list.append(split_exif[lenth][3])
	elif "1st" == split_exif[lenth][0] :
		_1st_list.append(split_exif[lenth][1:4])
		_1st_tag_list.append(split_exif[lenth][1])
		_1st_file_split.append(split_exif[lenth][2])
		_1st_value_list.append(split_exif[lenth][3])
	else :
		print ("!")

_0th_tag_list = list(set(_0th_tag_list))
Exif_tag_list = list(set(Exif_tag_list))
GPS_tag_list = list(set(GPS_tag_list))
Interop_tag_list = list(set(Interop_tag_list))
_1st_tag_list = list(set(_1st_tag_list))

_0th_value_list = list(set(_0th_value_list))
Exif_value_list = list(set(Exif_value_list))
GPS_value_list = list(set(GPS_value_list))
Interop_value_list = list(set(Interop_value_list))
_1st_value_list = list(set(_1st_value_list))

#ifd = _0th, Exif, GPS, Interop, _1st
#tag_name = tag_list / filename = file_split / value = value_list
#print (exif_list)
print ("\nParsing complete")

excel = Workbook()
sheet = excel.active
string = "Tag / Filename"
sheet.title = "Oth"
string = "Tag / Filename"
sheet['A1'] = string
sheet['A1'].alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
sheet.freeze_panes = 'A2'

for no_sheet in range (1, 5) :
	if no_sheet == 1 :
		exif_sheet = excel.create_sheet(title = 'Exif')
		exif_sheet['A1'] = string
		exif_sheet['A1'].alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
		exif_sheet.freeze_panes = 'A2'
	elif no_sheet == 2 :
		gps_sheet = excel.create_sheet(title = 'GPS')
		gps_sheet['A1'] = string
		gps_sheet['A1'].alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
		gps_sheet.freeze_panes = 'A2'
	elif no_sheet == 3 : 
		interop_sheet = excel.create_sheet(title = 'Interop')
		interop_sheet['A1'] = string
		interop_sheet['A1'].alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
		interop_sheet.freeze_panes = 'A2'
	else :
		_1st_sheet = excel.create_sheet(title = '1st')
		_1st_sheet['A1'] = string
		_1st_sheet['A1'].alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
		_1st_sheet.freeze_panes = 'A2'

# chr(65) = A
# => 0th
strlen = list()
for col in range(0, len(f_list)) :
	sheet[str(chr(col+66)) + "1"] = f_list[col]
	tempcell = sheet[str(chr(col+66)) + "1"]
	sheet.column_dimensions['A'].width = len(string)
	sheet.column_dimensions[str(chr(col+66))].width = len(f_list[col])
	tempcell.font = openpyxl.styles.Font(size=10)
	tempcell.alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
	tempcell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=openpyxl.styles.Color('e0cd66'))

for num in range (0, len(_0th_tag_list)) :
	sheet[str(chr(65)) + str(num+2)] = _0th_tag_list[num]
	tempvercell = sheet[str(chr(65)) + str(num+2)]
	tempvercell.font = openpyxl.styles.Font(size=10)
	tempvercell.alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
	tempvercell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=openpyxl.styles.Color('e3e3e3'))
	strlen.append(len(_0th_tag_list[num]))
	if num == len(_0th_tag_list)-1 :
		strlen.sort()
		sheet.column_dimensions[str(chr(65))].width = strlen[num]

strlen = list() #chr(66) = B #split_ exif - [0] : ifd / [1] : tag_name / [2] : filename / [3] : value
for col in range(0, len(f_list)) :
	for num in range (0, len(_0th_tag_list)) :
		for total in range(0, len(_0th_list)) :
			if (sheet[str(chr(66+col)) + "1"].value == _0th_list[total][1]) and (sheet[str(chr(65)) + str(num+2)].value == _0th_list[total][0]) :
				sheet[str(chr(66+col)) + str(num+2)] = str(_0th_list[total][2])
				tempvercell = sheet[str(chr(66+col)) + str(num+2)]
				tempvercell.font = openpyxl.styles.Font(size=10)
				tempvercell.alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
				#tempvercell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=openpyxl.styles.Color('e3e3e3'))
				#strlen.append(len(split_exif[total][3]))
#strlen.sort(reverse=True)
"""for col in range(0, len(f_list)) :
	for num in range (0, len(_0th_tag_list)) :
		sheet.column_dimensions[str(chr(66+col)) + str(num+2)].width = strlen[0]"""

# => Exif
for col in range(0, len(f_list)) :
	exif_sheet[str(chr(col+66)) + "1"] = f_list[col]
	tempcell = exif_sheet[str(chr(col+66)) + "1"]
	exif_sheet.column_dimensions['A'].width = len(string)
	exif_sheet.column_dimensions[str(chr(col+66))].width = len(f_list[col])
	tempcell.font = openpyxl.styles.Font(size=10)
	tempcell.alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
	tempcell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=openpyxl.styles.Color('e0cd66'))

for num in range (0, len(Exif_tag_list)) :
	exif_sheet[str(chr(65)) + str(num+2)] = Exif_tag_list[num]
	tempvercell = exif_sheet[str(chr(65)) + str(num+2)]
	tempvercell.font = openpyxl.styles.Font(size=10)
	tempvercell.alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
	tempvercell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=openpyxl.styles.Color('e3e3e3'))
	strlen.append(len(Exif_tag_list[num]))
	if num == len(Exif_tag_list)-1 :
		strlen.sort()
		exif_sheet.column_dimensions[str(chr(65))].width = strlen[num]

for col in range(0, len(f_list)) :
	for num in range (0, len(Exif_tag_list)) :
		for total in range(0, len(Exif_list)) :
			if (exif_sheet[str(chr(66+col)) + "1"].value == Exif_list[total][1]) and (exif_sheet[str(chr(65)) + str(num+2)].value == Exif_list[total][0]) :
				exif_sheet[str(chr(66+col)) + str(num+2)] = str(Exif_list[total][2])
				tempvercell = exif_sheet[str(chr(66+col)) + str(num+2)]
				tempvercell.font = openpyxl.styles.Font(size=10)
				tempvercell.alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')

# => GPS
strlen = list()
for col in range(0, len(f_list)) :
	gps_sheet[str(chr(col+66)) + "1"] = f_list[col]
	tempcell = gps_sheet[str(chr(col+66)) + "1"]
	gps_sheet.column_dimensions['A'].width = len(string)
	gps_sheet.column_dimensions[str(chr(col+66))].width = len(f_list[col])
	tempcell.font = openpyxl.styles.Font(size=10)
	tempcell.alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
	tempcell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=openpyxl.styles.Color('e0cd66'))

for num in range (0, len(GPS_tag_list)) :
	gps_sheet[str(chr(65)) + str(num+2)] = GPS_tag_list[num]
	tempvercell = gps_sheet[str(chr(65)) + str(num+2)]
	tempvercell.font = openpyxl.styles.Font(size=10)
	tempvercell.alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
	tempvercell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=openpyxl.styles.Color('e3e3e3'))
	strlen.append(len(GPS_tag_list[num]))
	if num == len(GPS_tag_list)-1 :
		strlen.sort()
		gps_sheet.column_dimensions[str(chr(65))].width = strlen[num]	

for col in range(0, len(f_list)) :
	for num in range (0, len(GPS_tag_list)) :
		for total in range(0, len(GPS_list)) :
			if (gps_sheet[str(chr(66+col)) + "1"].value == GPS_list[total][1]) and (gps_sheet[str(chr(65)) + str(num+2)].value == GPS_list[total][0]) :
				if GPS_list[total][0] == "GPSLatitude" :							
					latDeg = int(GPS_list[total][2].split(",")[0][2:]) / float(int(GPS_list[total][2].split(",")[1].split(" ")[1][:-1]))
					latMin = int(GPS_list[total][2].split(",")[1].split(" ")[1][:-1]) / float(int(GPS_list[total][2].split(",")[3].split(" ")[1][:-1]))
					latSec = int(GPS_list[total][2].split(",")[4].split(" ")[1][1:]) / float(int(GPS_list[total][2].split(",")[5].split(" ")[1][:-2]))
					Lat = (latDeg + (latMin + latSec / 60.0) / 60.0)
					GPS_list[total][2] = round(Lat, 6)
				if GPS_list[total][0] == "GPSLongitude" :					
					lonDeg = int(GPS_list[total][2].split(",")[0][2:]) / float(int(GPS_list[total][2].split(",")[1].split(" ")[1][:-1]))
					lonMin = int(GPS_list[total][2].split(",")[1].split(" ")[1][:-1]) / float(int(GPS_list[total][2].split(",")[3].split(" ")[1][:-1]))
					lonSec = int(GPS_list[total][2].split(",")[4].split(" ")[1][1:]) / float(int(GPS_list[total][2].split(",")[5].split(" ")[1][:-2]))
					Lon = (lonDeg + (lonMin + lonSec / 60.0) / 60.0)
					#GPS_list[total][2] = Lon
					GPS_list[total][2] = round(Lon, 6)
				if GPS_list[total][0] == "GPSTimeStamp" :							
					clock = round(int(GPS_list[total][2].split(",")[0][2:]) / int(GPS_list[total][2].split(",")[1].split(" ")[1][:-1]))
					minute = round(int(GPS_list[total][2].split(",")[1].split(" ")[1][:-1]) / float(int(GPS_list[total][2].split(",")[3].split(" ")[1][:-1])))
					second = round(int(GPS_list[total][2].split(",")[4].split(" ")[1][1:]) / float(int(GPS_list[total][2].split(",")[5].split(" ")[1][:-2])))
					time = str(clock) + ":" + str(minute) + ":" + str(second)
					GPS_list[total][2] = time
				if GPS_list[total][0] == "GPSAltitude" :					
					alti = int(GPS_list[total][2].split(",")[0][1:]) / float(int(GPS_list[total][2].split(",")[1].split(" ")[1][:-1]))
					GPS_list[total][2] = alti
				gps_sheet[str(chr(66+col)) + str(num+2)] = str(GPS_list[total][2])
				tempvercell = gps_sheet[str(chr(66+col)) + str(num+2)]
				tempvercell.font = openpyxl.styles.Font(size=10)
				tempvercell.alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')

# => Interop
strlen = list()
for col in range(0, len(f_list)) :
	interop_sheet[str(chr(col+66)) + "1"] = f_list[col]
	tempcell = interop_sheet[str(chr(col+66)) + "1"]
	interop_sheet.column_dimensions['A'].width = len(string)
	interop_sheet.column_dimensions[str(chr(col+66))].width = len(f_list[col])
	tempcell.font = openpyxl.styles.Font(size=10)
	tempcell.alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
	tempcell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=openpyxl.styles.Color('e0cd66'))

for num in range (0, len(Interop_tag_list)) :
	interop_sheet[str(chr(65)) + str(num+2)] = Interop_tag_list[num]
	tempvercell = interop_sheet[str(chr(65)) + str(num+2)]
	tempvercell.font = openpyxl.styles.Font(size=10)
	tempvercell.alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
	tempvercell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=openpyxl.styles.Color('e3e3e3'))
	strlen.append(len(Interop_tag_list[num]))
	if num == len(Interop_tag_list)-1 :
		strlen.sort()
		interop_sheet.column_dimensions[str(chr(65))].width = strlen[num]

for col in range(0, len(f_list)) :
	for num in range (0, len(Interop_tag_list)) :
		for total in range(0, len(Interop_list)) :
			if (interop_sheet[str(chr(66+col)) + "1"].value == Interop_list[total][1]) and (interop_sheet[str(chr(65)) + str(num+2)].value == Interop_list[total][0]) :
				interop_sheet[str(chr(66+col)) + str(num+2)] = str(Interop_list[total][2])
				tempvercell = interop_sheet[str(chr(66+col)) + str(num+2)]
				tempvercell.font = openpyxl.styles.Font(size=10)
				tempvercell.alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')	

# => 1st
strlen = list()
for col in range(0, len(f_list)) :
	_1st_sheet[str(chr(col+66)) + "1"] = f_list[col]
	tempcell = _1st_sheet[str(chr(col+66)) + "1"]
	_1st_sheet.column_dimensions['A'].width = len(string)
	_1st_sheet.column_dimensions[str(chr(col+66))].width = len(f_list[col])
	tempcell.font = openpyxl.styles.Font(size=10)
	tempcell.alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
	tempcell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=openpyxl.styles.Color('e0cd66'))

for num in range (0, len(_1st_tag_list)) :
	_1st_sheet[str(chr(65)) + str(num+2)] = _1st_tag_list[num]
	tempvercell = _1st_sheet[str(chr(65)) + str(num+2)]
	tempvercell.font = openpyxl.styles.Font(size=10)
	tempvercell.alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
	tempvercell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=openpyxl.styles.Color('e3e3e3'))
	strlen.append(len(_1st_tag_list[num]))
	if num == len(_1st_tag_list)-1 :
		strlen.sort()
		_1st_sheet.column_dimensions[str(chr(65))].width = strlen[num]

for col in range(0, len(f_list)) :
	for num in range (0, len(_1st_tag_list)) :
		for total in range(0, len(_1st_list)) :
			if (_1st_sheet[str(chr(66+col)) + "1"].value == _1st_list[total][1]) and (_1st_sheet[str(chr(65)) + str(num+2)].value == _1st_list[total][0]) :
				_1st_sheet[str(chr(66+col)) + str(num+2)] = str(_1st_list[total][2])
				tempvercell = _1st_sheet[str(chr(66+col)) + str(num+2)]
				tempvercell.font = openpyxl.styles.Font(size=10)
				tempvercell.alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
savename = "excel_parser_exif.xlsx"
excel.save(str(path) +"/" +str(savename))
print ("Finish the create Excel file for jpeg exif\n")
print ("save path : " + str(path) + "/" +str(savename))
